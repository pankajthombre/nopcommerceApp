import openpyxl
import openpyxl.styles import PatternFill

#created this utility for data driven TC, before this we created Test_data.xlsx file into TestData folder
def getRowcount(file,sheetname):
    workbook=openpyxl.Workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getcolcount(file,sheetname):
    workbook=openpyxl.Workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def readdata(file,sheetname,rownum,columnno):
    workbook=openpyxl.Workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,columnno).value #this will always return text(string) data

def writedata(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.Workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

def fillGreencolor(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

def fillreadcolor(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redFill=PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)